import os
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from supabase import create_client, Client
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

app = Flask(__name__)
# Vercel's root directory is different, so we ensure template/static paths are correct
app.template_folder = '../templates'
app.static_folder = '../static'
app.secret_key = os.getenv('SECRET_KEY', 'default_secret_key_change_me')

# Supabase Setup
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Auth Setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Simple User Model for Global Restriction
class User(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    if user_id == os.getenv('ADMIN_USERNAME', 'admin'):
        return User(user_id)
    return None

# Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Restriction: Check against configured admin credentials
        if username == os.getenv('ADMIN_USERNAME', 'admin') and password == os.getenv('ADMIN_PASSWORD', 'admin123'):
            user = User(username)
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid credentials. Access restricted.')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/admin')
@login_required
def admin():
    return render_template('admin.html')

@app.route('/verify', methods=['POST'])
@login_required
def verify():
    data = request.get_json()
    roll_number = data.get('roll_number', '').strip()
    
    if not roll_number:
        return jsonify({'status': 'error', 'message': 'Roll number is required'}), 400
    
    # Query Supabase
    response = supabase.table('students').select('*').eq('roll_number', roll_number).execute()
    
    if len(response.data) > 0:
        student = response.data[0]
        if student['is_scanned']:
            return jsonify({
                'status': 'duplicate',
                'message': f'Duplicate Entry ⚠️: {roll_number} already scanned.',
                'roll_number': roll_number
            })
        else:
            # Update to scanned
            supabase.table('students').update({
                'is_scanned': True,
                'scanned_at': datetime.now().isoformat()
            }).eq('roll_number', roll_number).execute()
            
            return jsonify({
                'status': 'authorized',
                'message': f'Authorized ✅: {roll_number} verified.',
                'roll_number': roll_number
            })
    else:
        return jsonify({
            'status': 'not_authorized',
            'message': f'Not Authorized ❌: {roll_number} not found in database.',
            'roll_number': roll_number
        })

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.xlsx'):
        return jsonify({'status': 'error', 'message': 'Invalid file format. Please upload .xlsx'}), 400
    
    try:
        wb = load_workbook(file)
        sheet = wb.active
        
        header = [cell.value for cell in sheet[1]]
        if 'roll_number' not in header:
            return jsonify({'status': 'error', 'message': 'Column "roll_number" not found in Excel'}), 400
        
        col_idx = header.index('roll_number') + 1
        roll_numbers = []
        for row in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row, column=col_idx).value
            if val:
                roll_numbers.append(str(val).strip())
        
        # Batch insert to Supabase (limit per request to avoid timeout)
        count = 0
        batch_size = 50
        for i in range(0, len(roll_numbers), batch_size):
            batch = [{'roll_number': r} for r in roll_numbers[i:i+batch_size]]
            try:
                # use upsert to skip duplicates if possible, or simple insert
                supabase.table('students').upsert(batch, on_conflict='roll_number').execute()
                count += len(batch)
            except Exception as e:
                print(f"Batch insert error: {e}")
        
        return jsonify({'status': 'success', 'message': f'Successfully processed {len(roll_numbers)} students.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/stats')
@login_required
def get_stats():
    # Fetch stats from Supabase
    total_res = supabase.table('students').select('roll_number', count='exact').execute()
    total = total_res.count if total_res.count is not None else 0
    
    scanned_res = supabase.table('students').select('roll_number', count='exact').eq('is_scanned', True).execute()
    scanned = scanned_res.count if scanned_res.count is not None else 0
    
    remaining = total - scanned
    
    recent_res = supabase.table('students').select('roll_number', 'scanned_at').eq('is_scanned', True).order('scanned_at', desc=True).limit(10).execute()
    recent = [{'roll_number': r['roll_number'], 'scanned_at': r['scanned_at']} for r in recent_res.data]
    
    return jsonify({
        'total': total,
        'scanned': scanned,
        'remaining': remaining,
        'recent': recent
    })

@app.route('/export')
@login_required
def export_data():
    recent_res = supabase.table('students').select('*').eq('is_scanned', True).execute()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Scanned Students"
    ws.append(['roll_number', 'is_scanned', 'scanned_at'])
    
    for student in recent_res.data:
        ws.append([student['roll_number'], student['is_scanned'], student['scanned_at']])
    
    export_path = '/tmp/scanned_students.xlsx' # Use /tmp for Vercel
    wb.save(export_path)
    return send_file(export_path, as_attachment=True)

if __name__ == '__main__':
    # Local fallback
    app.run(debug=True, port=5000)
